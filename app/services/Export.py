import os
import tempfile
import asyncio
from datetime import datetime
from typing import Dict, Any, List
from io import BytesIO
import requests
from PIL import Image as PILImage  # for fixing logo distortion
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.worksheet.filters import AutoFilter, FilterColumn
from bson import ObjectId
from app.models.Property import PropertyModel
from app.models.Operator import OperatorModel
from app.models.CompetitorProperty import CompetitorPropertyModel
from app.models.CompetitorComparison import CompetitorComparisonModel
from app.services.CompetitorComparison import CompetitorComparisonService
from app.services.Property import PropertyService
from app.helpers.AzureStorage import AzureBlobUploader
from app.helpers.ExcelFormatting import ExcelFormattingHelper
import re

# Global constants
LOGO_URL = "https://strategycuesstorage.blob.core.windows.net/strategy-cues-storage/strategy-cues-storage/7574c178258039f8.png"

class ExportService:
    def __init__(self):
        self.property_model = PropertyModel()
        self.azure_uploader = AzureBlobUploader()
        self.operator_model = OperatorModel()
        self.property_service = PropertyService()

    async def export_properties_to_excel(self, operator_id: str, filters: dict = None, property_ids: list = None) -> dict:
        """Export selected property fields to Excel with logo, custom header, and professional styling
        
        Args:
            operator_id: The operator ID
            filters: Optional filter dictionary
            property_ids: Optional list of property IDs to filter by
            
        Logic:
            - If both filters and property_ids: Apply filters first, then filter results to only include property_ids
            - If only property_ids: Return only those properties (no other filters)
            - If only filters: Apply filters only (current behavior)
            - If neither: Return all properties for operator
        """
        temp_file_path = None
        try:
            # Validate operator_id
            if not operator_id or not isinstance(operator_id, str):
                return {"success": False, "data": None, "error": "Invalid operator ID"}

            # Fetch operator name
            operator_data = await self.operator_model.get_operator({"_id": ObjectId(operator_id)})
            operator_name = operator_data.model_dump().get("name", str(operator_id))
            safe_operator_name = re.sub(r'[^a-zA-Z0-9_-]', '_', operator_name)

            # Build MongoDB query
            query = {"operator_id": operator_id}
            
            # Handle property_ids and filters logic
            if property_ids and filters:
                # Both provided: Apply filters first, then filter by property_ids in Python
                query = self._build_filter_query(filters)
            elif property_ids:
                # Only property_ids: Query only those properties
                # Convert property_ids to ObjectIds for MongoDB query
                try:
                    object_ids = [ObjectId(pid) for pid in property_ids if ObjectId.is_valid(pid)]
                    if object_ids:
                        query["_id"] = {"$in": object_ids}
                    else:
                        return {"success": False, "data": None, "error": "Invalid property IDs format"}
                except Exception as e:
                    return {"success": False, "data": None, "error": f"Invalid property IDs: {str(e)}"}
            elif filters:
                # Only filters: Apply filters only
                query = self._build_filter_query(filters)
            # else: Neither provided - query all properties for operator (default behavior)
            
            # Fetch properties with optimized projection (using query)
            properties = await self.property_model.get_properties_for_csv_optimized(query)
            if not properties:
                return {"success": False, "data": None, "error": f"No properties found for operator: {operator_id}"}
            
            # If both filters and property_ids were provided, filter results by property_ids
            if property_ids and filters:
                property_ids_set = set(property_ids)
                properties = [
                    prop for prop in properties 
                    if str(prop.get("_id", "")) in property_ids_set
                ]
                if not properties:
                    return {"success": False, "data": None, "error": f"No properties match both filters and provided property IDs"}

            # Create Excel workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Property Report"

            # === Define headers with listing name first and links at end ===
            link_columns = ["BookingUrl", "AirbnbUrl", "VRBOUrl"]
            all_headers = [
                "Listing_Name",  # Will be first
                "Area", "Room_Type", "Property_Type",
                # Occupancy
                "Occupancy This Month", "Occupancy Next Month", "Occupancy Same Time Last Year", "Occupancy Same Time Last Month",
                "Occupancy 7 Days", "Occupancy 30 Days",
                # Revenue (assuming RevPAR represents Revenue)
                "Revenue This Month", "Revenue Next Month",
                # ADR
                "ADR This Month", "ADR Next Month", "ADR Same Time Last Year", "ADR Same Time Last Month",
                # Pickup Occupancy
                "Pickup Occupancy 7 Days", "Pickup Occupancy 14 Days", "Pickup Occupancy 30 Days",
                # MPI
                "MPI This Month", "MPI Next Month", "MPI Last Year This Month",
                # Min Rate Threshold
                "Min Rate Threshold",
                # RevPAR
                "RevPAR Same Time Last Year", "RevPAR Same Time Last Month",
                # Reviews - Airbnb
                "Airbnb Total Reviews", "Airbnb Last Review Date", "Airbnb Last Review Score",
                # Reviews - Booking
                "Booking Total Reviews", "Booking Last Review Date", "Booking Last Review Score",
                # Booking.com Features
                "Genius", "Mobile", "Preferred",
                # Images
                "Booking_Image", "Airbnb_Image", "VRBO_Image",
                # Links at end
                "BookingUrl", "AirbnbUrl", "VRBOUrl"
            ]
            
            # Reorder headers using helper
            headers, listing_col_idx, link_start_col_idx = ExcelFormattingHelper.reorder_columns_with_listing_first(
                all_headers, link_columns
            )
            total_columns = len(headers)

            # === Header section ===
            # Use the helper function (header is now row 3-4, no logo at top)
            ExcelFormattingHelper.setup_sheet_header(ws, headers, operator_name, None, "Property Report")

            # === Populate data ===
            current_row = 3  # Start data from row 3 (after headers in row 1-2)
            for idx, prop in enumerate(properties, start=1):
                # Properties are now dicts, no need for model_dump()
                prop_dict = prop
                flat_prop = self._flatten_property_dict(prop_dict)
                flat_prop = self._convert_special_types(flat_prop)

                # Get URLs for links (will be added at end)
                booking_url = prop_dict.get("BookingUrl", "")
                airbnb_url = prop_dict.get("AirbnbUrl", "")
                vrbo_url = prop_dict.get("VRBOUrl", "")
                
                # Add listing name first (column 1) - left aligned and bold
                listing_name = prop_dict.get("Listing_Name", "")
                cell = ws.cell(row=current_row, column=1, value=listing_name)
                cell.font = Font(name="Arial", size=9, bold=True)
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
                ccc_border = Border(
                    left=Side(style='thin', color="CCCCCC"), 
                    right=Side(style='thin', color="CCCCCC"),
                    top=Side(style='thin', color="CCCCCC"), 
                    bottom=Side(style='thin', color="CCCCCC")
                )
                cell.border = ccc_border
                
                # Batch data preparation for better performance - using correct nested structure access
                row_data = [
                    prop_dict.get("Area", ""),
                    prop_dict.get("Room_Type", ""),
                    prop_dict.get("Property_Type", ""),
                    # Occupancy - using correct nested structure
                    prop_dict.get("Occupancy", {}).get("TM", ""),
                    prop_dict.get("Occupancy", {}).get("NM", ""),
                    prop_dict.get("STLY_Var", {}).get("Occ", ""),
                    prop_dict.get("STLM_Var", {}).get("Occ", ""),
                    prop_dict.get("Occupancy", {}).get("7_days", ""),
                    prop_dict.get("Occupancy", {}).get("30_days", ""),  # Using 30_days instead of 14_days as per JSON
                    # Revenue - using correct nested structure
                    prop_dict.get("RevPAR", {}).get("TM", ""),
                    prop_dict.get("RevPAR", {}).get("NM", ""),
                    # ADR - using correct nested structure
                    prop_dict.get("ADR", {}).get("TM", ""),
                    prop_dict.get("ADR", {}).get("NM", ""),
                    prop_dict.get("STLY_Var", {}).get("ADR", ""),
                    prop_dict.get("STLM_Var", {}).get("ADR", ""),
                    # Pickup Occupancy - using correct nested structure
                    prop_dict.get("Pick_Up_Occ", {}).get("7_Days", ""),
                    prop_dict.get("Pick_Up_Occ", {}).get("14_Days", ""),
                    prop_dict.get("Pick_Up_Occ", {}).get("30_Days", ""),
                    # MPI - using correct nested structure
                    prop_dict.get("MPI", {}).get("TM", ""),
                    prop_dict.get("MPI", {}).get("NM", ""),
                    prop_dict.get("MPI", {}).get("LYTM", ""),
                    # Min Rate Threshold
                    prop_dict.get("Min_Rate_Threshold", ""),
                    # RevPAR - using correct nested structure
                    prop_dict.get("STLY_Var", {}).get("RevPAR", ""),
                    prop_dict.get("STLM_Var", {}).get("RevPAR", ""),
                    # Reviews - Airbnb - using correct nested structure
                    prop_dict.get("Reviews", {}).get("Airbnb", {}).get("Total_Rev", ""),
                    prop_dict.get("Reviews", {}).get("Airbnb", {}).get("Last_Review_Date", ""),
                    prop_dict.get("Reviews", {}).get("Airbnb", {}).get("Last_Rev_Score", ""),
                    # Reviews - Booking - using correct nested structure
                    prop_dict.get("Reviews", {}).get("Booking", {}).get("Total_Rev", ""),
                    prop_dict.get("Reviews", {}).get("Booking", {}).get("Last_Review_Date", ""),
                    prop_dict.get("Reviews", {}).get("Booking", {}).get("Last_Rev_Score", ""),
                    # Booking.com Features - Genius, Mobile, Preferred
                    self._format_booking_com_value(prop_dict.get("BookingCom", {}), "Genius"),
                    self._format_booking_com_value(prop_dict.get("BookingCom", {}), "Mobile"),
                    self._format_booking_com_value(prop_dict.get("BookingCom", {}), "Pref")
                ]
                
                # Batch write data with alignment (columns 2 onwards, skipping listing name) - right aligned
                for col_idx, value in enumerate(row_data, start=2):
                    ExcelFormattingHelper.set_cell_value_with_alignment(ws, current_row, col_idx, value, is_listing_name_col=False)
                
                # Images - Convert to hyperlinks (before links)
                photos_data = prop_dict.get("Photos", {})
                booking_photo = self._get_first_photo(photos_data, "booking")
                airbnb_photo = self._get_first_photo(photos_data, "airbnb")
                vrbo_photo = self._get_first_photo(photos_data, "vrbo")
                
                # Calculate image column position (before links)
                image_col = total_columns - 3  # 3 columns before links
                image_urls = [booking_photo, airbnb_photo, vrbo_photo]
                for i, photo_url in enumerate(image_urls):
                    if photo_url:
                        ExcelFormattingHelper.add_hyperlink_to_cell(ws, current_row, image_col + i, photo_url, None, is_listing_name_col=False)
                
                # Add link hyperlinks at the end
                link_urls = [booking_url, airbnb_url, vrbo_url]
                for i, url in enumerate(link_urls):
                    if url:
                        ExcelFormattingHelper.add_hyperlink_to_cell(ws, current_row, link_start_col_idx + i, url, None, is_listing_name_col=False)

                # Apply row formatting (white background, #ccc borders only for cells with data, alignment)
                ExcelFormattingHelper.apply_row_coloring(ws, current_row, total_columns, is_listing_name_col=True)
                current_row += 1

            # Calculate last data row before adding logo
            last_data_row = current_row - 1
            
            # Freeze listing name column (freeze at row 3, after header rows 1-2)
            ExcelFormattingHelper.freeze_listing_name_column(ws, listing_col_idx, 3)
            
            # Calculate column indices for URLs and Images for proper auto-sizing
            # Image columns are 3 columns before links (Booking_Image, Airbnb_Image, VRBO_Image)
            image_col_start = total_columns - 3
            image_columns = [image_col_start, image_col_start + 1, image_col_start + 2]
            # URL columns are at the end (BookingUrl, AirbnbUrl, VRBOUrl)
            url_columns = [link_start_col_idx, link_start_col_idx + 1, link_start_col_idx + 2]
            
            # Auto-fit column widths (before adding logo, so it doesn't affect sizing)
            ExcelFormattingHelper.auto_fit_columns(ws, total_columns, last_data_row, url_columns, image_columns, listing_col_idx)
            
            # Add column filters (only up to last_data_row, before logo)
            ExcelFormattingHelper.add_column_filters(ws, total_columns, last_data_row)
            
            # Add logo to bottom (after filters, so it doesn't interfere)
            ExcelFormattingHelper.add_logo_to_bottom(ws, LOGO_URL, total_columns, last_data_row)

            # Save Excel with operator name
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_file:
                temp_file_path = temp_file.name
                wb.save(temp_file_path)

            final_file_path = os.path.join(tempfile.gettempdir(), f"{safe_operator_name}.xlsx")
            os.rename(temp_file_path, final_file_path)

            # Apply filters by reopening the file (header starts at row 1)
            final_file_path = ExcelFormattingHelper.apply_filters_to_excel_file(final_file_path, header_row=1)

            # Upload to Azure Blob
            folder_name = f"exports/{operator_id}"
            file_url = self.azure_uploader.upload_excel_file_to_azure_blob(
                final_file_path,
                folder_name=folder_name,
                file_name=safe_operator_name,
                file_type=".xlsx"
            )

            # Cleanup
            if final_file_path and os.path.exists(final_file_path):
                os.unlink(final_file_path)
            
            # Cleanup temp logo file if it exists
            temp_logo_path = os.path.join(tempfile.gettempdir(), "temp_logo.png")
            if os.path.exists(temp_logo_path):
                os.unlink(temp_logo_path)

            if not file_url:
                return {"success": False, "data": None, "error": "Failed to upload Excel to Azure"}

            return {"success": True, "data": {"file_url": file_url, "expiry_time": 3600}}

        except Exception as e:
            if temp_file_path and os.path.exists(temp_file_path):
                os.unlink(temp_file_path)
            return {"success": False, "data": None, "error": str(e)}

    # ===== Keep all your other functions intact =====
    def _flatten_property_dict(self, d: Dict[str, Any], parent_key: str = '', sep: str = '_') -> Dict[str, Any]:
        items = []
        for k, v in d.items():
            new_key = f"{parent_key}{sep}{k}" if parent_key else k
            if isinstance(v, dict):
                if k == "Photos":
                    items.append((new_key, v))  # Keep Photos dict for later
                else:
                    items.extend(self._flatten_property_dict(v, new_key, sep=sep).items())
            elif isinstance(v, list):
                items.append((new_key, ', '.join(str(i) for i in v) if v else ""))
            else:
                # Handle enum values - extract the enum name/value
                if hasattr(v, 'value'):  # Enum with value
                    items.append((new_key, v.value if v.value is not None else ""))
                elif hasattr(v, 'name'):  # Enum with name
                    items.append((new_key, v.name if v.name is not None else ""))
                else:
                    items.append((new_key, v if v is not None else ""))
        return dict(items)

    def _convert_special_types(self, flat_prop: Dict[str, Any]) -> Dict[str, Any]:
        for key, value in flat_prop.items():
            if isinstance(value, datetime):
                flat_prop[key] = value.strftime('%Y-%m-%d %H:%M:%S')
            elif isinstance(value, ObjectId):
                flat_prop[key] = str(value)
            elif value is None:
                flat_prop[key] = ""
        return flat_prop

    def _get_first_photo(self, photos_dict: Dict[str, Any], platform: str) -> str:
        if not photos_dict:
            return ""
        platform_photos = photos_dict.get(platform, [])
        if platform_photos and isinstance(platform_photos, list):
            first_photo = platform_photos[0]
            if isinstance(first_photo, dict):
                return first_photo.get('url', "")
        return ""
    
    def _format_booking_com_value(self, booking_com: Dict[str, Any], field: str) -> str:
        """Return fixed percentage overrides for Booking.com participation flags."""
        if not booking_com:
            return ""
        
        value = booking_com.get(field, "")
        if value == "Yes":
            percentage_mapping = {
                "Genius": "10%",
                "Mobile": "10%",
                "Pref": "5%"
            }
            return percentage_mapping.get(field, "")
        if value == "No":
            return "0"
        return ""

    async def export_content_cues_to_excel(self, operator_id: str) -> dict:
        """Export Content Cues competitor analysis to Excel with two simplified sheets"""
        temp_file_path = None
        try:
            # Validate operator_id
            if not operator_id or not isinstance(operator_id, str):
                return {"success": False, "data": None, "error": "Invalid operator ID"}

            # Fetch all data in parallel for maximum performance
            competitor_model = CompetitorPropertyModel()
            
            # Parallel data fetching - optimized (removed comparison_task)
            operator_task = self.operator_model.get_operator({"_id": ObjectId(operator_id)})
            properties_task = self.property_model.get_properties_for_content_cues_optimized({"operator_id": operator_id})
            
            # Wait for properties first to get property IDs for AI analysis
            operator_data, properties = await asyncio.gather(operator_task, properties_task)
            
            # Extract property IDs for targeted AI analysis fetching
            property_ids = [str(prop.get("_id", "")) for prop in properties if prop.get("_id")]
            
            # Fetch AI analysis data for specific properties (optimized)
            comparison_model = CompetitorComparisonModel()
            ai_analysis_data = await comparison_model.get_comparisons_for_ai_analysis_by_property_ids(property_ids)
            
            # Extract operator name
            operator_name = operator_data.model_dump().get("name", str(operator_id))
            safe_operator_name = re.sub(r'[^a-zA-Z0-9_-]', '_', operator_name)

            if not properties:
                return {"success": False, "data": None, "error": f"No properties found for operator: {operator_id}"}

            # Extract competitor IDs from properties data (optimized - simplified with set comprehension)
            competitor_ids = {
                competitor_id 
                for prop in properties 
                for competitor_id in prop.get("competitorIds", []) or []
                if competitor_id
            }
            
            # Fetch only the required competitor properties with optimized counts
            competitor_properties = []
            if competitor_ids:
                competitor_properties = await competitor_model.get_competitors_with_counts_optimized(list(competitor_ids))

            # Amenities data is now pre-calculated in database aggregation

            # Create Excel workbook
            wb = Workbook()
            # Remove default sheet and create our custom sheets
            wb.remove(wb.active)
            
            # Create three sheets
            overview_sheet = wb.create_sheet("Overview")
            ai_analysis_booking_sheet = wb.create_sheet("AI Analysis - Booking")
            ai_analysis_airbnb_sheet = wb.create_sheet("AI Analysis - Airbnb")

            # === CREATE ALL SHEETS ===
            self._create_overview_sheet(overview_sheet, properties, competitor_properties, operator_name, operator_id, LOGO_URL)
            self._create_ai_analysis_booking_sheet(ai_analysis_booking_sheet, properties, ai_analysis_data, operator_name, operator_id, LOGO_URL)
            self._create_ai_analysis_airbnb_sheet(ai_analysis_airbnb_sheet, properties, ai_analysis_data, operator_name, operator_id, LOGO_URL)

            # Save temp Excel file
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_file:
                temp_file_path = temp_file.name
                wb.save(temp_file_path)

            # Rename temp file with desired filename
            final_file_path = os.path.join(tempfile.gettempdir(), f"ContentCues_{safe_operator_name}.xlsx")
            os.rename(temp_file_path, final_file_path)

            # Apply filters by reopening the file (header starts at row 1)
            final_file_path = ExcelFormattingHelper.apply_filters_to_excel_file(final_file_path, header_row=1)

            # Upload to Azure Blob asynchronously
            folder_name = f"exports/{operator_id}"
            file_url = await asyncio.get_event_loop().run_in_executor(
                None,
                self.azure_uploader.upload_excel_file_to_azure_blob,
                final_file_path,
                folder_name,
                f"ContentCues_{safe_operator_name}",
                ".xlsx"
            )
            
            # Cleanup
            if final_file_path and os.path.exists(final_file_path):
                os.unlink(final_file_path)
            
            # Cleanup temp logo file if it exists
            temp_logo_path = os.path.join(tempfile.gettempdir(), "temp_logo.png")
            if os.path.exists(temp_logo_path):
                os.unlink(temp_logo_path)

            if not file_url:
                return {"success": False, "data": None, "error": "Failed to upload Excel to Azure"}

            return {"success": True, "data": {"file_url": file_url, "expiry_time": 3600}}
        except Exception as e:
            if temp_file_path and os.path.exists(temp_file_path):
                os.unlink(temp_file_path)
            return {"success": False, "data": None, "error": str(e)}


    def _has_caption(self, photo):
        """Check if photo has a valid caption (null-safe)"""
        try:
            caption = photo.get("caption")
            return caption is not None and str(caption).strip() != ""
        except (AttributeError, TypeError):
            return False

    def _create_overview_sheet(self, ws, properties, competitor_properties, operator_name, operator_id, logo_url):
        """Create Overview sheet with property and competitor comparison data"""
        # Define headers for overview with competitor data
        headers = [
            "Property_Name", "Booking_Link", "Airbnb_Link",
            "Airbnb_Total_Photos", "Airbnb_Photos_Without_Caption", "Airbnb_Amenities_Count", "Airbnb_Reviews",
            "Booking_Total_Photos", "Booking_Photos_Without_Caption", "Booking_Amenities_Count", "Booking_Reviews",
            "Competitor_Name", "Competitor_Link", 
            "Competitor_Booking_Photos", "Competitor_Airbnb_Photos",
            "Competitor_Booking_Reviews", "Competitor_Airbnb_Reviews",
            "Competitor_Booking_Amenities", "Competitor_Airbnb_Amenities"
        ]
        
        ExcelFormattingHelper.setup_sheet_header(ws, headers, operator_name, None, "Property Overview")
        
        # Create a map of competitor properties by their IDs for easy lookup
        competitor_map = {}
        for comp_prop in competitor_properties:
            comp_id = str(comp_prop["_id"])
            competitor_map[comp_id] = comp_prop
        
        # Pre-process property data for better performance
        property_data_cache = {}
        for prop in properties:
            # Properties are now dicts, no need for model_dump()
            prop_dict = prop
            property_id = str(prop.get("_id", ""))
            
            # Use pre-calculated counts from database aggregation
            property_data_cache[property_id] = {
                "listing_name": prop_dict.get("Listing_Name", ""),
                "booking_url": prop_dict.get("BookingUrl", ""),
                "airbnb_url": prop_dict.get("AirbnbUrl", ""),
                
                # Use database-aggregated photo counts
                "airbnb_total_photos": prop_dict.get("total_photos_airbnb", 0),
                "booking_total_photos": prop_dict.get("total_photos_booking", 0),
                "airbnb_photos_without_caption": prop_dict.get("total_photos_airbnb", 0) - prop_dict.get("photos_with_captions_airbnb", 0),
                "booking_photos_without_caption": prop_dict.get("total_photos_booking", 0) - prop_dict.get("photos_with_captions_booking", 0),
                
                # Use database-aggregated review counts
                "airbnb_reviews": prop_dict.get("reviews_airbnb_total", 0),
                "booking_reviews": prop_dict.get("reviews_booking_total", 0),
                
                # Use database-aggregated amenity counts
                "airbnb_amenities_count": prop_dict.get("amenities_airbnb_count", 0),
                "booking_amenities_count": prop_dict.get("amenities_booking_count", 0)
            }
        
        # Populate data starting at row 3
        current_row = 3  # Start data from row 3 (after headers in row 1-2)
        row_counter = 0
        for idx, prop in enumerate(properties, start=1):
            property_id = str(prop.get("_id", ""))
            prop_data = property_data_cache[property_id]
            
            # Get competitor IDs for this property
            property_competitor_ids = prop.get("competitorIds", [])
            
            if not property_competitor_ids:
                row_counter += 1
                # No competitors case - show just property data
                # Add property name first (column 1) - left aligned
                ExcelFormattingHelper.set_cell_value_with_alignment(ws, current_row, 1, prop_data["listing_name"], is_listing_name_col=True)
                
                # Add hyperlinks for Booking and Airbnb URLs - right aligned
                if prop_data["booking_url"]:
                    ExcelFormattingHelper.add_hyperlink_to_cell(ws, current_row, 2, prop_data["booking_url"], None, is_listing_name_col=False)
                else:
                    ExcelFormattingHelper.set_cell_value_with_alignment(ws, current_row, 2, "", is_listing_name_col=False)
                
                if prop_data["airbnb_url"]:
                    ExcelFormattingHelper.add_hyperlink_to_cell(ws, current_row, 3, prop_data["airbnb_url"], None, is_listing_name_col=False)
                else:
                    ExcelFormattingHelper.set_cell_value_with_alignment(ws, current_row, 3, "", is_listing_name_col=False)
                
                # Add the rest of the data using pre-calculated values - batch operation
                row_data = [
                    prop_data["airbnb_total_photos"],
                    prop_data["airbnb_photos_without_caption"],
                    prop_data["airbnb_amenities_count"],
                    prop_data["airbnb_reviews"],
                    prop_data["booking_total_photos"],
                    prop_data["booking_photos_without_caption"],
                    prop_data["booking_amenities_count"],
                    prop_data["booking_reviews"],
                    "No Competitors",
                    "",
                    0,  # Competitor_Booking_Photos
                    0,  # Competitor_Airbnb_Photos
                    0,  # Competitor_Booking_Reviews
                    0,  # Competitor_Airbnb_Reviews
                    0,  # Competitor_Booking_Amenities
                    0   # Competitor_Airbnb_Amenities
                ]
                
                # Batch write to Excel with proper alignment (right aligned)
                for col_idx, value in enumerate(row_data, start=4):
                    ExcelFormattingHelper.set_cell_value_with_alignment(ws, current_row, col_idx, value, is_listing_name_col=False)
                
                # Apply row formatting (white background, #ccc borders only for cells with data, alignment)
                ExcelFormattingHelper.apply_row_coloring(ws, current_row, len(headers), is_listing_name_col=True)
                current_row += 1
            else:
                # Show property data with each competitor
                for comp_idx, competitor_id in enumerate(property_competitor_ids):
                    row_counter += 1
                    if competitor_id not in competitor_map:
                        continue  # Skip if competitor not found in our fetched data
                    
                    comp_prop_data = competitor_map[competitor_id]
                    competitor_name = comp_prop_data.get("propertyName", "Unknown Competitor")
                    competitor_link = comp_prop_data.get("bookingLink", "") or comp_prop_data.get("airbnbLink", "")
                    
                    # Use pre-calculated counts from database aggregation
                    comp_booking_photos = comp_prop_data.get("competitor_booking_photos_count", 0)
                    comp_airbnb_photos = comp_prop_data.get("competitor_airbnb_photos_count", 0)
                    comp_booking_reviews = comp_prop_data.get("competitor_booking_reviews_count", 0)
                    comp_airbnb_reviews = comp_prop_data.get("competitor_airbnb_reviews_count", 0)
                    comp_booking_amenities_count = comp_prop_data.get("competitor_booking_amenities_count", 0)
                    comp_airbnb_amenities_count = comp_prop_data.get("competitor_airbnb_amenities_count", 0)
                    
                    # Add property name first (column 1) - left aligned
                    ExcelFormattingHelper.set_cell_value_with_alignment(ws, current_row, 1, prop_data["listing_name"], is_listing_name_col=True)
                    
                    # Add hyperlinks for Booking and Airbnb URLs - right aligned
                    if prop_data["booking_url"]:
                        ExcelFormattingHelper.add_hyperlink_to_cell(ws, current_row, 2, prop_data["booking_url"], None, is_listing_name_col=False)
                    else:
                        ExcelFormattingHelper.set_cell_value_with_alignment(ws, current_row, 2, "", is_listing_name_col=False)
                    
                    if prop_data["airbnb_url"]:
                        ExcelFormattingHelper.add_hyperlink_to_cell(ws, current_row, 3, prop_data["airbnb_url"], None, is_listing_name_col=False)
                    else:
                        ExcelFormattingHelper.set_cell_value_with_alignment(ws, current_row, 3, "", is_listing_name_col=False)
                    
                    # Add property data using pre-calculated values - right aligned
                    ExcelFormattingHelper.set_cell_value_with_alignment(ws, current_row, 4, prop_data["airbnb_total_photos"], is_listing_name_col=False)
                    ExcelFormattingHelper.set_cell_value_with_alignment(ws, current_row, 5, prop_data["airbnb_photos_without_caption"], is_listing_name_col=False)
                    ExcelFormattingHelper.set_cell_value_with_alignment(ws, current_row, 6, prop_data["airbnb_amenities_count"], is_listing_name_col=False)
                    ExcelFormattingHelper.set_cell_value_with_alignment(ws, current_row, 7, prop_data["airbnb_reviews"], is_listing_name_col=False)
                    ExcelFormattingHelper.set_cell_value_with_alignment(ws, current_row, 8, prop_data["booking_total_photos"], is_listing_name_col=False)
                    ExcelFormattingHelper.set_cell_value_with_alignment(ws, current_row, 9, prop_data["booking_photos_without_caption"], is_listing_name_col=False)
                    ExcelFormattingHelper.set_cell_value_with_alignment(ws, current_row, 10, prop_data["booking_amenities_count"], is_listing_name_col=False)
                    ExcelFormattingHelper.set_cell_value_with_alignment(ws, current_row, 11, prop_data["booking_reviews"], is_listing_name_col=False)
                    
                    # Add competitor data - right aligned
                    ExcelFormattingHelper.set_cell_value_with_alignment(ws, current_row, 12, competitor_name, is_listing_name_col=False)
                    
                    # Add competitor hyperlink if available - right aligned
                    if competitor_link:
                        ExcelFormattingHelper.add_hyperlink_to_cell(ws, current_row, 13, competitor_link, None, is_listing_name_col=False)
                    else:
                        ExcelFormattingHelper.set_cell_value_with_alignment(ws, current_row, 13, "", is_listing_name_col=False)
                    
                    # Competitor data with separate Booking and Airbnb columns - right aligned
                    ExcelFormattingHelper.set_cell_value_with_alignment(ws, current_row, 14, comp_booking_photos, is_listing_name_col=False)      # Competitor_Booking_Photos
                    ExcelFormattingHelper.set_cell_value_with_alignment(ws, current_row, 15, comp_airbnb_photos, is_listing_name_col=False)      # Competitor_Airbnb_Photos
                    ExcelFormattingHelper.set_cell_value_with_alignment(ws, current_row, 16, comp_booking_reviews, is_listing_name_col=False)    # Competitor_Booking_Reviews
                    ExcelFormattingHelper.set_cell_value_with_alignment(ws, current_row, 17, comp_airbnb_reviews, is_listing_name_col=False)      # Competitor_Airbnb_Reviews
                    ExcelFormattingHelper.set_cell_value_with_alignment(ws, current_row, 18, comp_booking_amenities_count, is_listing_name_col=False)  # Competitor_Booking_Amenities
                    ExcelFormattingHelper.set_cell_value_with_alignment(ws, current_row, 19, comp_airbnb_amenities_count, is_listing_name_col=False)   # Competitor_Airbnb_Amenities
                    
                    # Apply row formatting (white background, #ccc borders only for cells with data, alignment)
                    ExcelFormattingHelper.apply_row_coloring(ws, current_row, len(headers), is_listing_name_col=True)
                    current_row += 1
        
        # Calculate last data row before adding logo
        last_data_row = current_row - 1
        
        # Freeze listing name column (freeze at row 3, after header rows 1-2)
        ExcelFormattingHelper.freeze_listing_name_column(ws, 1, 3)
        
        # Calculate column indices for URLs in overview sheet
        # Property_Name (1), Booking_Link (2), Airbnb_Link (3), ..., Competitor_Link (13)
        # Find Competitor_Link column index
        competitor_link_col = None
        for idx, header in enumerate(headers, start=1):
            if "Competitor_Link" in header or "Competitor Link" in header:
                competitor_link_col = idx
                break
        url_columns = [2, 3]  # Booking_Link and Airbnb_Link columns
        if competitor_link_col:
            url_columns.append(competitor_link_col)
        
        # Auto-fit column widths (before adding logo, so it doesn't affect sizing)
        ExcelFormattingHelper.auto_fit_columns(ws, len(headers), last_data_row, url_columns=url_columns, listing_name_col=1)

        # Add column filters (only up to last_data_row, before logo)
        ExcelFormattingHelper.add_column_filters(ws, len(headers), last_data_row)
        
        # Add logo to bottom (after filters, so it doesn't interfere)
        ExcelFormattingHelper.add_logo_to_bottom(ws, logo_url, len(headers), last_data_row)

    def _add_column_filters(self, ws, total_columns):
        """Add column filters to the data range"""
        try:
            # Get the actual last row with data
            last_row = ws.max_row
            if last_row >= 5:  # Make sure we have at least the header row
                # Apply auto filter to the entire data range
                filter_range = f"A5:{get_column_letter(total_columns)}{last_row}"
                ws.auto_filter.ref = filter_range
                
        except Exception as e:
            # print(f"Auto filter error: {e}")  # Debug info
            pass

    def _apply_filters_to_excel_file(self, file_path: str, header_row: int = 1) -> str:
        """
        Apply Excel auto-filters to all sheets in an existing Excel file.
        Ensures each column gets its own filter dropdown at the header row.
        """
        try:
            wb = load_workbook(file_path)
            for ws in wb.worksheets:
                max_row, max_col = ws.max_row, ws.max_column
                if max_row > header_row and max_col > 0:
                    # Apply auto filter across full data range starting from header row
                    filter_range = f"A{header_row}:{get_column_letter(max_col)}{max_row}"
                    
                    # Clear existing auto filter and recreate with proper column filters
                    ws.auto_filter = AutoFilter()
                    ws.auto_filter.ref = filter_range
                    
                    # Add filter columns for each column (filters appear in header row)
                    for col_idx in range(1, max_col + 1):
                        filter_col = FilterColumn(colId=col_idx - 1)  # colId is 0-based
                        ws.auto_filter.filterColumn.append(filter_col)

            wb.save(file_path)
            wb.close()
            return file_path

        except Exception as e:
            # print(f"⚠️ Error applying filters: {e}")
            return file_path

    def _set_cell_value_with_alignment(self, ws, row, col, value):
        """Set cell value with consistent left alignment"""
        cell = ws.cell(row=row, column=col, value=value)
        cell.alignment = Alignment(
            horizontal="left", 
            vertical="center",
            wrap_text=True
        )
        return cell

    def _add_hyperlink_to_cell(self, ws, row, col, url, text=None):
        """Add hyperlink to a cell with actual URL as text"""
        try:
            cell = ws.cell(row=row, column=col)
            # Use the actual URL as the text if no custom text provided
            display_text = text if text else url
            cell.value = display_text
            cell.hyperlink = Hyperlink(ref=cell.coordinate, target=url)
            cell.font = Font(color="0000FF", underline="single")
            # Apply left alignment to hyperlink cells too
            cell.alignment = Alignment(
                horizontal="left", 
                vertical="center",
                wrap_text=True
            )
        except Exception as e:
            # If hyperlink fails, just add the URL as text with alignment
            self._set_cell_value_with_alignment(ws, row, col, url)

    def _setup_sheet_header(self, ws, headers, operator_name, logo_url, sheet_title):
        """Setup common header for all sheets with logo and formatting"""
        total_columns = len(headers)
        last_column = get_column_letter(total_columns)
        
        # Row 1-4: Merge cells for logo on the left (A1:A4)
        ws.merge_cells("A1:A4")
        logo_cell = ws["A1"]
        logo_cell.alignment = Alignment(horizontal="center", vertical="center")
        no_border = Border(left=Side(style=None), right=Side(style=None), 
                         top=Side(style=None), bottom=Side(style=None))
        for row in ws["A1:A4"]:
            for cell in row:
                cell.border = no_border
                cell.fill = PatternFill(fill_type=None)
        
        # Insert logo
        temp_logo_path = None
        if logo_url:
            try:
                resp = requests.get(logo_url)
                pil_img = PILImage.open(BytesIO(resp.content))
                max_width, max_height = 190, 75
                pil_img.thumbnail((max_width, max_height), PILImage.LANCZOS)
                temp_logo_path = os.path.join(tempfile.gettempdir(), "temp_logo.png")
                pil_img.save(temp_logo_path)
                img = Image(temp_logo_path)
                img.anchor = "A1"
                ws.add_image(img, "A1")
                logo_cell = ws["A1"]
                logo_cell.alignment = Alignment(
                    horizontal="center", 
                    vertical="center"
                )
            except:
                pass

        # Row 1-4: Report title with operator name (B1:last_column)
        ws.merge_cells(f"B1:{last_column}4")
        title_cell = ws["B1"]
        title_cell.value = f"{sheet_title} - {operator_name}"
        title_cell.font = Font(size=16, bold=True)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        title_cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        title_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                           top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Row 5: Column headers
        for col_idx, header_value in enumerate(headers, start=1):
            cell = ws.cell(row=5, column=col_idx)
            cell.value = header_value
            cell.font = Font(bold=True, size=12)
            # No background color for headers
            cell.alignment = Alignment(
                horizontal="center", 
                vertical="center",
                wrap_text=True
            )
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                           top=Side(style='thin'), bottom=Side(style='thin'))

    def _apply_row_coloring(self, ws, row_index, total_columns):
        """Apply alternating row coloring and cell padding"""
        fill_color = "FFF9E5" if row_index % 2 == 0 else "FFFFFF" 
        for col_idx in range(1, total_columns + 1):
            cell = ws.cell(row=row_index, column=col_idx)
            cell.fill = PatternFill(
                start_color=fill_color, end_color=fill_color, fill_type="solid"
            )
            # Add consistent padding and alignment to each cell
            cell.alignment = Alignment(
                horizontal="left", 
                vertical="center",
                wrap_text=True
            )

    def _auto_fit_columns(self, ws, total_columns):
        """Auto-fit column widths similar to export_properties_to_excel"""
        from openpyxl.cell.cell import MergedCell
        
        for col_idx in range(1, total_columns + 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            
            for row_idx in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                try:
                    # Skip merged cells and handle potential non-string values
                    if not isinstance(cell, MergedCell) and cell.value is not None:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except (TypeError, AttributeError):
                    pass  
            adjusted_width = max_length + 1 if max_length > 0 else 10  # Minimum width of 10
            ws.column_dimensions[column_letter].width = adjusted_width

    def _create_ai_analysis_booking_sheet(self, ws, properties, ai_analysis_data, operator_name, operator_id, logo_url):
        """Create AI Analysis - Booking sheet with Booking.com specific AI-generated insights"""
        try:
            # Define column headers for AI Analysis - Booking
            headers = [
                "Property Name",
                "Photo Suggestion",
                "Competitors Positive Points", 
                "Competitors Negative Points",
                "Suggested Changes - Competitor Reviews",
                "My Positive Points",
                "My Negative Points", 
                "Suggested Changes - My Reviews",
                "Guests Wishlist - Competitor",
                "Guests Wishlist - My Property",
                "Suggested Amenities"
            ]
            
            # Set up header (no logo at top)
            ExcelFormattingHelper.setup_sheet_header(ws, headers, operator_name, None, "AI Analysis - Booking")
            
            # Create a map of property_id to AI analysis data
            ai_analysis_map = {}
            for analysis in ai_analysis_data:
                property_id = analysis.get("propertyId", "")
                if property_id:
                    ai_analysis_map[property_id] = analysis
            
            # Populate data starting from row 3
            current_row = 3
            for idx, prop in enumerate(properties, start=1):
                property_id = str(prop.get("_id", ""))
                property_name = prop.get("Listing_Name", "")
                
                # Get AI analysis data for this property
                analysis = ai_analysis_map.get(property_id, {})
                
                # Photo suggestions - Booking specific
                ai_photo_analysis_booking = analysis.get("aiPhotoAnalysisBooking", {})
                recommendations = ai_photo_analysis_booking.get("summary", {}).get("recommendations", [])
                photo_suggestions = "\n".join(str(rec) for rec in recommendations if rec) if recommendations else ""
                
                # Competitor insights - Booking specific
                review_analysis_competitor_booking = analysis.get("reviewAnalysisCompetitorBooking", [])
                
                # Competitor positive points
                competitor_positive = "\n".join([
                    f"{insight.get('title', '')}\n{insight.get('description', '')}" 
                    for insight in review_analysis_competitor_booking
                    if insight.get("type") == "LOVEDTOHAVE"
                ])
                
                # Competitor negative points
                competitor_negative = "\n".join([
                    f"{insight.get('title', '')}\n{insight.get('description', '')}" 
                    for insight in review_analysis_competitor_booking
                    if insight.get("type") == "DIDNTLIKE"
                ])
                
                # Competitor suggestions - Booking specific (only titles)
                competitor_suggestions = "\n".join([
                    suggestion.get('title', '') 
                    for suggestion in analysis.get("reviewSuggestionsBasedOnCompetitorBooking", [])
                    if suggestion.get('title')
                ])
                
                # Own property insights - Booking specific
                review_analysis_own_booking = analysis.get("reviewAnalysisOwnBooking", [])
                
                # Own positive points
                own_positive = "\n".join([
                    f"{insight.get('title', '')}\n{insight.get('description', '')}" 
                    for insight in review_analysis_own_booking
                    if insight.get("type") == "LOVEDTOHAVE"
                ])
                
                # Own negative points
                own_negative = "\n".join([
                    f"{insight.get('title', '')}\n{insight.get('description', '')}" 
                    for insight in review_analysis_own_booking
                    if insight.get("type") == "DIDNTLIKE"
                ])
                
                # Own suggestions - Booking specific (only titles)
                own_suggestions = "\n".join([
                    suggestion.get('title', '') 
                    for suggestion in analysis.get("reviewSuggestionsBasedOnOwnBooking", [])
                    if suggestion.get('title')
                ])
                
                # Guests wishlist - Competitor (Booking)
                wishlist_competitor = "\n".join([
                    f"{insight.get('title', '')}\n{insight.get('description', '')}" 
                    for insight in review_analysis_competitor_booking
                    if insight.get("type") == "WISHTOHAVE"
                ])
                
                # Guests wishlist - My Property (Booking)
                wishlist_my_property = "\n".join([
                    f"{insight.get('title', '')}\n{insight.get('description', '')}" 
                    for insight in review_analysis_own_booking
                    if insight.get("type") == "WISHTOHAVE"
                ])
                
                # Suggested amenities - from conversionBoostersBooking.missing
                conversion_boosters_booking = analysis.get("conversionBoostersBooking", {})
                missing_amenities = conversion_boosters_booking.get("missing", [])
                suggested_amenities = "\n".join([
                    amenity.get('label', '') 
                    for amenity in missing_amenities
                    if amenity.get('label')
                ])
                
                # Populate row data
                row_data = [
                    property_name,
                    photo_suggestions,
                    competitor_positive,
                    competitor_negative,
                    competitor_suggestions,
                    own_positive,
                    own_negative,
                    own_suggestions,
                    wishlist_competitor,
                    wishlist_my_property,
                    suggested_amenities
                ]
                
                # Write data first
                for col_idx, value in enumerate(row_data, start=1):
                    ws.cell(row=current_row, column=col_idx, value=value)
                
                # Apply row formatting
                ExcelFormattingHelper.apply_row_coloring(ws, current_row, len(headers), is_listing_name_col=True)
                
                # Override alignment for all columns to be left-aligned, vertically centered with text wrapping
                max_lines = 1  # Track maximum lines in this row for dynamic height
                for col_idx in range(1, len(headers) + 1):
                    cell = ws.cell(row=current_row, column=col_idx)
                    value = cell.value
                    
                    # Calculate approximate number of lines based on content
                    if value:
                        value_str = str(value)
                        # Estimate lines: count newlines + estimate wrapping based on content length
                        newline_count = value_str.count('\n')
                        # Rough estimate: ~50 characters per line for wrapped text (adjust based on column width)
                        estimated_wrapped_lines = max(1, len(value_str) // 50)
                        total_lines = max(newline_count + 1, estimated_wrapped_lines)
                        max_lines = max(max_lines, total_lines)
                    
                    cell.alignment = Alignment(
                        horizontal="left", 
                        vertical="center",
                        wrap_text=True
                    )
                    cell.font = Font(name="Arial", size=9)
                
                # Set dynamic row height (reduced padding: ~12 points per line, minimum 15)
                row_height = max(15, max_lines * 12)
                ws.row_dimensions[current_row].height = row_height
                
                current_row += 1
            
            # Calculate last data row before adding logo
            last_data_row = current_row - 1
            
            # Freeze listing name column
            ExcelFormattingHelper.freeze_listing_name_column(ws, 1, 3)
            
            # Auto-fit columns
            ExcelFormattingHelper.auto_fit_columns(ws, len(headers), last_data_row, listing_name_col=1)
            
            # Add column filters
            ExcelFormattingHelper.add_column_filters(ws, len(headers), last_data_row)
            
            # Add logo to bottom
            ExcelFormattingHelper.add_logo_to_bottom(ws, logo_url, len(headers), last_data_row)
            
        except Exception as e:
            print(f"❌ Error creating AI Analysis - Booking sheet: {e}")

    def _create_ai_analysis_airbnb_sheet(self, ws, properties, ai_analysis_data, operator_name, operator_id, logo_url):
        """Create AI Analysis - Airbnb sheet with Airbnb specific AI-generated insights"""
        try:
            # Define column headers for AI Analysis - Airbnb
            headers = [
                "Property Name",
                "Photo Suggestion",
                "Competitors Positive Points", 
                "Competitors Negative Points",
                "Suggested Changes - Competitor Reviews",
                "My Positive Points",
                "My Negative Points", 
                "Suggested Changes - My Reviews",
                "Guests Wishlist - Competitor",
                "Guests Wishlist - My Property",
                "Suggested Amenities"
            ]
            
            # Set up header (no logo at top)
            ExcelFormattingHelper.setup_sheet_header(ws, headers, operator_name, None, "AI Analysis - Airbnb")
            
            # Create a map of property_id to AI analysis data
            ai_analysis_map = {}
            for analysis in ai_analysis_data:
                property_id = analysis.get("propertyId", "")
                if property_id:
                    ai_analysis_map[property_id] = analysis
            
            # Populate data starting from row 3
            current_row = 3
            for idx, prop in enumerate(properties, start=1):
                property_id = str(prop.get("_id", ""))
                property_name = prop.get("Listing_Name", "")
                
                # Get AI analysis data for this property
                analysis = ai_analysis_map.get(property_id, {})
                
                # Photo suggestions - Airbnb specific
                ai_photo_analysis_airbnb = analysis.get("aiPhotoAnalysisAirbnb", {})
                recommendations = ai_photo_analysis_airbnb.get("summary", {}).get("recommendations", [])
                photo_suggestions = "\n".join(str(rec) for rec in recommendations if rec) if recommendations else ""
                
                # Competitor insights - Airbnb specific
                review_analysis_competitor_airbnb = analysis.get("reviewAnalysisCompetitorAirbnb", [])
                
                # Competitor positive points
                competitor_positive = "\n".join([
                    f"{insight.get('title', '')}\n{insight.get('description', '')}" 
                    for insight in review_analysis_competitor_airbnb
                    if insight.get("type") == "LOVEDTOHAVE"
                ])
                
                # Competitor negative points
                competitor_negative = "\n".join([
                    f"{insight.get('title', '')}\n{insight.get('description', '')}" 
                    for insight in review_analysis_competitor_airbnb
                    if insight.get("type") == "DIDNTLIKE"
                ])
                
                # Competitor suggestions - Airbnb specific (only titles)
                competitor_suggestions = "\n".join([
                    suggestion.get('title', '') 
                    for suggestion in analysis.get("reviewSuggestionsBasedOnCompetitorAirbnb", [])
                    if suggestion.get('title')
                ])
                
                # Own property insights - Airbnb specific
                review_analysis_own_airbnb = analysis.get("reviewAnalysisOwnAirbnb", [])
                
                # Own positive points
                own_positive = "\n".join([
                    f"{insight.get('title', '')}\n{insight.get('description', '')}" 
                    for insight in review_analysis_own_airbnb
                    if insight.get("type") == "LOVEDTOHAVE"
                ])
                
                # Own negative points
                own_negative = "\n".join([
                    f"{insight.get('title', '')}\n{insight.get('description', '')}" 
                    for insight in review_analysis_own_airbnb
                    if insight.get("type") == "DIDNTLIKE"
                ])
                
                # Own suggestions - Airbnb specific (only titles)
                own_suggestions = "\n".join([
                    suggestion.get('title', '') 
                    for suggestion in analysis.get("reviewSuggestionsBasedOnOwnAirbnb", [])
                    if suggestion.get('title')
                ])
                
                # Guests wishlist - Competitor (Airbnb)
                wishlist_competitor = "\n".join([
                    f"{insight.get('title', '')}\n{insight.get('description', '')}" 
                    for insight in review_analysis_competitor_airbnb
                    if insight.get("type") == "WISHTOHAVE"
                ])
                
                # Guests wishlist - My Property (Airbnb)
                wishlist_my_property = "\n".join([
                    f"{insight.get('title', '')}\n{insight.get('description', '')}" 
                    for insight in review_analysis_own_airbnb
                    if insight.get("type") == "WISHTOHAVE"
                ])
                
                # Suggested amenities - from conversionBoostersAirbnb.missing
                conversion_boosters_airbnb = analysis.get("conversionBoostersAirbnb", {})
                missing_amenities = conversion_boosters_airbnb.get("missing", [])
                suggested_amenities = "\n".join([
                    amenity.get('label', '') 
                    for amenity in missing_amenities
                    if amenity.get('label')
                ])
                
                # Populate row data
                row_data = [
                    property_name,
                    photo_suggestions,
                    competitor_positive,
                    competitor_negative,
                    competitor_suggestions,
                    own_positive,
                    own_negative,
                    own_suggestions,
                    wishlist_competitor,
                    wishlist_my_property,
                    suggested_amenities
                ]
                
                # Write data first
                for col_idx, value in enumerate(row_data, start=1):
                    ws.cell(row=current_row, column=col_idx, value=value)
                
                # Apply row formatting
                ExcelFormattingHelper.apply_row_coloring(ws, current_row, len(headers), is_listing_name_col=True)
                
                # Override alignment for all columns to be left-aligned, vertically centered with text wrapping
                max_lines = 1  # Track maximum lines in this row for dynamic height
                for col_idx in range(1, len(headers) + 1):
                    cell = ws.cell(row=current_row, column=col_idx)
                    value = cell.value
                    
                    # Calculate approximate number of lines based on content
                    if value:
                        value_str = str(value)
                        # Estimate lines: count newlines + estimate wrapping based on content length
                        newline_count = value_str.count('\n')
                        # Rough estimate: ~50 characters per line for wrapped text (adjust based on column width)
                        estimated_wrapped_lines = max(1, len(value_str) // 50)
                        total_lines = max(newline_count + 1, estimated_wrapped_lines)
                        max_lines = max(max_lines, total_lines)
                    
                    cell.alignment = Alignment(
                        horizontal="left", 
                        vertical="center",
                        wrap_text=True
                    )
                    cell.font = Font(name="Arial", size=9)
                
                # Set dynamic row height (reduced padding: ~12 points per line, minimum 15)
                row_height = max(15, max_lines * 12)
                ws.row_dimensions[current_row].height = row_height
                
                current_row += 1
            
            # Calculate last data row before adding logo
            last_data_row = current_row - 1
            
            # Freeze listing name column
            ExcelFormattingHelper.freeze_listing_name_column(ws, 1, 3)
            
            # Auto-fit columns
            ExcelFormattingHelper.auto_fit_columns(ws, len(headers), last_data_row, listing_name_col=1)
            
            # Add column filters
            ExcelFormattingHelper.add_column_filters(ws, len(headers), last_data_row)
            
            # Add logo to bottom
            ExcelFormattingHelper.add_logo_to_bottom(ws, logo_url, len(headers), last_data_row)
            
        except Exception as e:
            print(f"❌ Error creating AI Analysis - Airbnb sheet: {e}")

    def _build_filter_query(self, filters: dict) -> dict:
        """Build MongoDB query from filters using PropertyService filter logic"""
        query = {}
        
        # Basic Filters
        if filters.get("operator_id"):
            query["operator_id"] = filters["operator_id"]
        if filters.get("area"):
            query["Area"] = {"$regex": f"^{filters['area']}$", "$options": "i"}
        if filters.get("room_type"):
            query["Room_Type"] = {"$regex": f"^{filters['room_type']}$", "$options": "i"}

        # ADR Range
        if filters.get("adr_range") and (filters["adr_range"]["min"] is not None or filters["adr_range"]["max"] is not None):
            query["ADR.TM"] = self.property_service._build_range_query(
                filters["adr_range"]["min"],
                filters["adr_range"]["max"],
                as_string=False
            )

        # RevPAR Range
        if filters.get("revpar_range") and (filters["revpar_range"]["min"] is not None or filters["revpar_range"]["max"] is not None):
            query["RevPAR.TM"] = self.property_service._build_range_query(
                filters["revpar_range"]["min"],
                filters["revpar_range"]["max"],
                as_string=False
            )

        # MPI Range
        if filters.get("mpi_range") and (filters["mpi_range"]["min"] is not None or filters["mpi_range"]["max"] is not None):
            query["MPI"] = self.property_service._build_range_query(
                self.property_service._strip_percentage(filters["mpi_range"]["min"]),
                self.property_service._strip_percentage(filters["mpi_range"]["max"]),
                as_string=False
            )

        # Min Rate Threshold
        if filters.get("min_rate_threshold") and (filters["min_rate_threshold"]["min"] is not None or filters["min_rate_threshold"]["max"] is not None):
            query["Min_Rate_Threshold"] = self.property_service._build_range_query(
                self.property_service._strip_percentage(filters["min_rate_threshold"]["min"]),
                self.property_service._strip_percentage(filters["min_rate_threshold"]["max"]),
                as_string=False
            )

        # Occupancy Filters
        if filters.get("occupancy"):
            self.property_service._add_occupancy_filters(query, filters["occupancy"])
        if filters.get("pickup"):
            self.property_service._add_pickup_filters(query, filters["pickup"])

        # Performance Filters
        if filters.get("stly_var"):
            self.property_service._add_stly_filters(query, filters["stly_var"])
        if filters.get("stlm_var"):
            self.property_service._add_stlm_filters(query, filters["stlm_var"])

        # Platform Features
        if filters.get("booking_features"):
            self.property_service._add_booking_feature_filters(query, filters["booking_features"])
        if filters.get("airbnb_features"):
            self.property_service._add_airbnb_feature_filters(query, filters["airbnb_features"])
        if filters.get("vrbo_features"):
            self.property_service._add_vrbo_feature_filters(query, filters["vrbo_features"])

        # Review Filters
        if filters.get("reviews"):
            self.property_service._add_review_filters(query, filters["reviews"])

        return query

